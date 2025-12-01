"""
Utilidades para exportar compras a PDF y Excel
"""
try:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False
    # Dummy values para evitar errores de importación
    A4 = None
    colors = None
    cm = None
    SimpleDocTemplate = None
    Table = None
    TableStyle = None
    Paragraph = None
    Spacer = None
    PageBreak = None
    getSampleStyleSheet = None
    ParagraphStyle = None
    TA_CENTER = None
    TA_RIGHT = None
    TA_LEFT = None

from datetime import datetime
from decimal import Decimal
from typing import List, Dict, Any
import io

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    Workbook = None
    Font = None
    Alignment = None
    PatternFill = None
    Border = None
    Side = None
    get_column_letter = None


def format_decimal(value) -> str:
    """Formatea un decimal a string con 2 decimales"""
    if value is None:
        return "0.00"
    return f"{float(value):.2f}"


def format_date(date_value) -> str:
    """Formatea una fecha a string"""
    if isinstance(date_value, str):
        return date_value
    if hasattr(date_value, 'strftime'):
        return date_value.strftime("%d/%m/%Y")
    return str(date_value)


def generate_purchase_pdf(purchase_data: Dict[str, Any], output_path: str = None) -> bytes:
    """
    Genera un PDF de una compra individual
    
    Args:
        purchase_data: Diccionario con los datos de la compra
        output_path: Ruta opcional para guardar el archivo. Si es None, retorna bytes
    
    Returns:
        bytes: Contenido del PDF
    
    Raises:
        ImportError: Si reportlab no está instalado
    """
    if not REPORTLAB_AVAILABLE:
        raise ImportError("reportlab no está instalado. Instálalo con: pip install reportlab")
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(output_path or buffer, pagesize=A4)
    story = []
    
    styles = getSampleStyleSheet()
    
    # Estilos personalizados
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#1e40af'),
        spaceAfter=30,
        alignment=TA_CENTER
    )
    
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=14,
        textColor=colors.HexColor('#1e40af'),
        spaceAfter=12,
        spaceBefore=12
    )
    
    # Título
    story.append(Paragraph("ORDEN DE COMPRA", title_style))
    
    # Información de la compra
    purchase_number = purchase_data.get('purchase_number', 'N/A')
    purchase_date = format_date(purchase_data.get('purchase_date'))
    supplier = purchase_data.get('supplier', {})
    supplier_name = supplier.get('name', 'N/A') if isinstance(supplier, dict) else 'N/A'
    
    info_data = [
        ['Número de Compra:', purchase_number],
        ['Fecha:', purchase_date],
        ['Proveedor:', supplier_name],
        ['Estado:', purchase_data.get('status', 'N/A').upper()],
    ]
    
    if purchase_data.get('reference_number'):
        info_data.append(['Referencia:', purchase_data.get('reference_number')])
    
    info_table = Table(info_data, colWidths=[5*cm, 10*cm])
    info_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('ALIGN', (1, 0), (1, -1), 'LEFT'),
    ]))
    story.append(info_table)
    
    story.append(Spacer(1, 0.5*cm))
    
    # Items de la compra
    story.append(Paragraph("DETALLE DE ITEMS", heading_style))
    
    items = purchase_data.get('items', [])
    if items:
        item_headers = ['Descripción', 'Cantidad', 'Precio Unit.', 'IVA %', 'Subtotal']
        item_rows = [item_headers]
        
        for item in items:
            item_rows.append([
                item.get('description', ''),
                format_decimal(item.get('quantity', 0)),
                f"€{format_decimal(item.get('unit_price', 0))}",
                f"{format_decimal(item.get('tax_rate', 0))}%",
                f"€{format_decimal(item.get('subtotal', 0))}",
            ])
        
        item_table = Table(item_rows, colWidths=[7*cm, 2*cm, 2.5*cm, 2*cm, 2.5*cm])
        item_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3b82f6')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
            ('ALIGN', (2, 0), (-1, -1), 'RIGHT'),
            ('ALIGN', (3, 0), (-1, -1), 'CENTER'),
            ('ALIGN', (4, 0), (-1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f9fafb')]),
        ]))
        story.append(item_table)
    else:
        story.append(Paragraph("No hay items registrados", styles['Normal']))
    
    # Totales
    story.append(Spacer(1, 0.5*cm))
    totals_data = [
        ['Subtotal:', f"€{format_decimal(purchase_data.get('subtotal', 0))}"],
        ['IVA:', f"€{format_decimal(purchase_data.get('tax', 0))}"],
        ['TOTAL:', f"€{format_decimal(purchase_data.get('total', 0))}"],
    ]
    
    totals_table = Table(totals_data, colWidths=[10*cm, 5*cm])
    totals_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (0, 2), (1, 2), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('FONTSIZE', (0, 2), (1, 2), 12),
        ('BACKGROUND', (0, 2), (1, 2), colors.HexColor('#dbeafe')),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
    ]))
    story.append(totals_table)
    
    # Notas
    if purchase_data.get('notes'):
        story.append(Spacer(1, 1*cm))
        story.append(Paragraph("NOTAS", heading_style))
        notes_para = Paragraph(purchase_data['notes'], styles['Normal'])
        story.append(notes_para)
    
    # Construir PDF
    doc.build(story)
    
    if output_path:
        return None
    
    buffer.seek(0)
    return buffer.read()


def generate_purchases_list_pdf(purchases: List[Dict[str, Any]], output_path: str = None) -> bytes:
    """
    Genera un PDF con una lista de compras
    
    Args:
        purchases: Lista de diccionarios con datos de compras
        output_path: Ruta opcional para guardar el archivo
    
    Returns:
        bytes: Contenido del PDF
    
    Raises:
        ImportError: Si reportlab no está instalado
    """
    if not REPORTLAB_AVAILABLE:
        raise ImportError("reportlab no está instalado. Instálalo con: pip install reportlab")
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(output_path or buffer, pagesize=A4, topMargin=2*cm)
    story = []
    
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#1e40af'),
        spaceAfter=20,
        alignment=TA_CENTER
    )
    
    # Título
    story.append(Paragraph("LISTADO DE COMPRAS", title_style))
    story.append(Paragraph(f"Generado el {datetime.now().strftime('%d/%m/%Y %H:%M')}", styles['Normal']))
    story.append(Spacer(1, 0.5*cm))
    
    # Tabla de compras
    headers = ['Número', 'Fecha', 'Proveedor', 'Estado', 'Subtotal', 'IVA', 'Total']
    rows = [headers]
    
    for purchase in purchases:
        supplier = purchase.get('supplier', {})
        supplier_name = supplier.get('name', 'N/A') if isinstance(supplier, dict) else 'N/A'
        
        rows.append([
            purchase.get('purchase_number', 'N/A'),
            format_date(purchase.get('purchase_date')),
            supplier_name[:30],  # Truncar si es muy largo
            purchase.get('status', 'N/A').upper(),
            f"€{format_decimal(purchase.get('subtotal', 0))}",
            f"€{format_decimal(purchase.get('tax', 0))}",
            f"€{format_decimal(purchase.get('total', 0))}",
        ])
    
    # Agregar totales
    total_subtotal = sum(Decimal(str(p.get('subtotal', 0))) for p in purchases)
    total_tax = sum(Decimal(str(p.get('tax', 0))) for p in purchases)
    total_total = sum(Decimal(str(p.get('total', 0))) for p in purchases)
    
    rows.append([
        '', '', '', 'TOTALES:',
        f"€{format_decimal(total_subtotal)}",
        f"€{format_decimal(total_tax)}",
        f"€{format_decimal(total_total)}",
    ])
    
    table = Table(rows, colWidths=[3*cm, 2.5*cm, 4*cm, 2*cm, 2.5*cm, 2*cm, 2.5*cm])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3b82f6')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('ALIGN', (4, 0), (-1, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('FONTSIZE', (0, -1), (-1, -1), 9),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('GRID', (0, 0), (-1, -2), 0.5, colors.grey),
        ('LINEBELOW', (0, -1), (-1, -1), 2, colors.HexColor('#1e40af')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#f9fafb')]),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#dbeafe')),
    ]))
    story.append(table)
    
    doc.build(story)
    
    if output_path:
        return None
    
    buffer.seek(0)
    return buffer.read()


def generate_purchases_excel(purchases: List[Dict[str, Any]], output_path: str = None) -> bytes:
    """
    Genera un archivo Excel con una lista de compras
    
    Args:
        purchases: Lista de diccionarios con datos de compras
        output_path: Ruta opcional para guardar el archivo
    
    Returns:
        bytes: Contenido del Excel
    
    Raises:
        ImportError: Si openpyxl no está instalado
    """
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl no está instalado. Instálalo con: pip install openpyxl")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Compras"
    
    # Estilos
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    right_align = Alignment(horizontal='right', vertical='center')
    
    # Encabezados
    headers = ['Número', 'Fecha', 'Proveedor', 'Estado', 'Subtotal', 'IVA', 'Total', 'Referencia']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border
    
    # Datos
    for row_num, purchase in enumerate(purchases, 2):
        supplier = purchase.get('supplier', {})
        supplier_name = supplier.get('name', 'N/A') if isinstance(supplier, dict) else 'N/A'
        
        ws.cell(row=row_num, column=1, value=purchase.get('purchase_number', 'N/A')).border = border
        ws.cell(row=row_num, column=2, value=format_date(purchase.get('purchase_date'))).border = border
        ws.cell(row=row_num, column=3, value=supplier_name).border = border
        ws.cell(row=row_num, column=4, value=purchase.get('status', 'N/A').upper()).border = border
        ws.cell(row=row_num, column=5, value=float(purchase.get('subtotal', 0))).border = border
        ws.cell(row=row_num, column=6, value=float(purchase.get('tax', 0))).border = border
        ws.cell(row=row_num, column=7, value=float(purchase.get('total', 0))).border = border
        ws.cell(row=row_num, column=8, value=purchase.get('reference_number', '')).border = border
        
        # Formato numérico para columnas de dinero
        for col in [5, 6, 7]:
            cell = ws.cell(row=row_num, column=col)
            cell.number_format = '#,##0.00'
            cell.alignment = right_align
    
    # Totales
    total_row = len(purchases) + 2
    ws.cell(row=total_row, column=1, value="TOTALES:").font = Font(bold=True)
    ws.cell(row=total_row, column=4, value="").border = border
    
    total_subtotal = sum(Decimal(str(p.get('subtotal', 0))) for p in purchases)
    total_tax = sum(Decimal(str(p.get('tax', 0))) for p in purchases)
    total_total = sum(Decimal(str(p.get('total', 0))) for p in purchases)
    
    ws.cell(row=total_row, column=5, value=float(total_subtotal)).font = Font(bold=True)
    ws.cell(row=total_row, column=5).number_format = '#,##0.00'
    ws.cell(row=total_row, column=5).alignment = right_align
    ws.cell(row=total_row, column=5).border = border
    
    ws.cell(row=total_row, column=6, value=float(total_tax)).font = Font(bold=True)
    ws.cell(row=total_row, column=6).number_format = '#,##0.00'
    ws.cell(row=total_row, column=6).alignment = right_align
    ws.cell(row=total_row, column=6).border = border
    
    ws.cell(row=total_row, column=7, value=float(total_total)).font = Font(bold=True)
    ws.cell(row=total_row, column=7).number_format = '#,##0.00'
    ws.cell(row=total_row, column=7).alignment = right_align
    ws.cell(row=total_row, column=7).border = border
    
    # Ajustar ancho de columnas
    column_widths = [15, 12, 25, 12, 12, 12, 12, 15]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = width
    
    # Guardar
    if output_path:
        wb.save(output_path)
        return None
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()

